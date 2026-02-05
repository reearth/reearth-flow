#!/usr/bin/env python3
import openpyxl
import csv
import sys
from io import StringIO

def excel_to_csv(file_path):
    """Convert all sheets in an Excel file to CSV format and print them."""
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # Process each sheet
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Print sheet header
            print(f"\n{'='*80}")
            print(f"Sheet: {sheet_name}")
            print(f"{'='*80}\n")

            # Create CSV output
            output = StringIO()
            writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

            # Get all rows
            rows_written = 0
            for row in sheet.iter_rows(values_only=True):
                # Convert None to empty string and handle all types
                cleaned_row = []
                for cell in row:
                    if cell is None:
                        cleaned_row.append('')
                    else:
                        cleaned_row.append(str(cell))

                writer.writerow(cleaned_row)
                rows_written += 1

            # Print the CSV content
            csv_content = output.getvalue()
            print(csv_content)

            # Print summary
            print(f"\n(Total rows: {rows_written})")

        print(f"\n{'='*80}")
        print(f"Total sheets processed: {len(wb.sheetnames)}")
        print(f"{'='*80}\n")

    except FileNotFoundError:
        print(f"Error: File not found: {file_path}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Error processing Excel file: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python excel_to_csv.py <excel_file_path>")
        sys.exit(1)

    excel_file = sys.argv[1]
    excel_to_csv(excel_file)
